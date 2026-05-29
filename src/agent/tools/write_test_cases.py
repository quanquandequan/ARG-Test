"""Test case generator tool — produces an Excel file from a requirements document.

Workflow (orchestrated by the Agent):
  1. Agent calls ``knowledge_search`` to retrieve existing test case samples
     from the knowledge base (for format/style reference).
  2. Agent calls ``write_test_cases`` with the requirement text and the KB
     samples.  This tool uses the injected LLM to generate structured test
     cases and then writes them to an Excel file.
  3. The tool returns the absolute path of the saved file so the Agent can
     relay it to the user.
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from src.agent.base_tool import BaseTool
from src.core.logging import get_logger
from src.llm.base import BaseLLM
from src.llm.types import Message

logger = get_logger(__name__)

# ── Default Excel column schema ──────────────────────────────────────────────
# Used when no column names can be inferred from kb_samples.
_DEFAULT_COLUMNS = [
    "用例编号",
    "所属模块",
    "用例标题",
    "前置条件",
    "测试步骤",
    "预期结果",
    "优先级",
    "用例类型",
    "备注",
]

_PRIORITY_VALUES = {"P0", "P1", "P2", "高", "中", "低"}
_TYPE_VALUES = {"正向", "反向", "边界", "异常", "功能", "兼容性"}

# ── LLM prompts ───────────────────────────────────────────────────────────────

_SYSTEM_PROMPT = """\
你是一名资深软件测试工程师，擅长根据需求文档编写完整的测试用例。
你的任务是将需求文档转换为结构化的测试用例列表，以 JSON 数组格式输出。

输出格式要求：
- 只输出 JSON 数组，不加任何 Markdown 标记或解释文字
- 每个测试用例是一个 JSON 对象，字段固定如下：
  {
    "title": "用例标题（简洁描述测试点）",
    "module": "所属功能模块",
    "precondition": "前置条件（无则填'无'）",
    "steps": "测试步骤（多步用换行分隔，每步以序号开头）",
    "expected": "预期结果（明确、可验证）",
    "priority": "优先级（P0/P1/P2）",
    "type": "用例类型（正向/反向/边界/异常）",
    "notes": "备注（无则留空）"
  }
- 覆盖正向路径、反向路径（无效输入/越权/异常）、边界值三类场景
- P0 = 核心主流程，P1 = 重要功能，P2 = 边缘场景
"""

_USER_TEMPLATE = """\
{kb_section}需求文档内容：
{requirement}

模块名称：{module}

请为上述需求生成完整的测试用例，输出 JSON 数组。
"""

_KB_SECTION_TEMPLATE = """\
以下是知识库中现有的测试用例样本，请参考其描述风格、粒度和术语，生成风格一致的新用例：

{samples}

---
"""


class WriteTestCasesTool(BaseTool):
    """Generate test cases from a requirements text and save to Excel.

    The tool is LLM-powered: it calls the injected language model to produce
    structured test cases, then writes them to an ``.xlsx`` file via openpyxl.

    Typical Agent usage:
      1. ``knowledge_search`` → retrieves existing test case samples for format reference
      2. ``write_test_cases`` → generates new cases and saves the Excel file

    All generation parameters (prompts, columns, output dir, temperature) are
    read from ``configs/default.yaml`` → ``test_generator``.  If a key is
    missing, the hardcoded default below is used.
    """

    # Prevent pytest from treating this as a test suite
    __test__ = False

    def __init__(
        self,
        llm: BaseLLM,
        output_dir: str | None = None,
        temperature: float | None = None,
        max_tokens: int | None = None,
        system_prompt: str | None = None,
        user_template: str | None = None,
        columns: list[dict] | None = None,
    ):
        from src.core.config import get_config

        cfg = get_config().get("test_generator", {})

        self._llm = llm
        self._default_output_dir = Path(
            output_dir or cfg.get("output_dir", "./outputs/test_cases")
        )
        self._temperature = (
            temperature if temperature is not None
            else float(cfg.get("temperature", 0.3))
        )
        self._max_tokens = (
            max_tokens if max_tokens is not None
            else int(cfg.get("max_tokens", 8192))
        )
        self._system_prompt = (
            system_prompt or cfg.get("system_prompt", "") or _SYSTEM_PROMPT
        )
        self._user_template = (
            user_template or cfg.get("user_template", "") or _USER_TEMPLATE
        )

        # Build column schema from explicit arg, config, or default
        columns_src = columns or cfg.get("columns")
        if columns_src and isinstance(columns_src, list):
            self._columns: list[tuple[str, int]] = [
                (c["name"], int(c.get("width", 15))) for c in columns_src
            ]
        else:
            self._columns = [(name, 15) for name in _DEFAULT_COLUMNS]

    @property
    def name(self) -> str:
        return "write_test_cases"

    @property
    def description(self) -> str:
        return (
            "根据需求文档和知识库现有用例格式，自动生成完整测试用例并保存为 Excel 文件。"
            "调用前请先使用 knowledge_search 获取现有测试用例样本以对齐格式和风格。"
            "返回生成的 Excel 文件路径和用例数量。"
        )

    @property
    def parameters(self) -> dict:
        return {
            "type": "object",
            "properties": {
                "requirement": {
                    "type": "string",
                    "description": "需求文档的完整文本内容",
                },
                "kb_samples": {
                    "type": "string",
                    "description": "从 knowledge_search 获取的现有测试用例样本（格式参考，可选）",
                },
                "module": {
                    "type": "string",
                    "description": "功能模块名称，用于用例分组和文件命名（可选）",
                },
                "output_dir": {
                    "type": "string",
                    "description": f"Excel 输出目录，默认为 {self._default_output_dir}",
                },
            },
            "required": ["requirement"],
        }

    async def execute(
        self,
        requirement: str = "",
        kb_samples: str = "",
        module: str = "",
        output_dir: str = "",
        **kwargs,
    ) -> str:
        if not requirement or not requirement.strip():
            return "错误：请提供需求文档内容。"

        module = module.strip() or "通用"
        out_dir = Path(output_dir.strip()) if output_dir.strip() else self._default_output_dir
        out_dir.mkdir(parents=True, exist_ok=True)

        # 1. Build LLM prompt
        kb_section = (
            _KB_SECTION_TEMPLATE.format(samples=kb_samples.strip())
            if kb_samples.strip()
            else ""
        )
        user_content = _USER_TEMPLATE.format(
            kb_section=kb_section,
            requirement=requirement.strip(),
            module=module,
        )
        messages = [
            Message(role="system", content=self._system_prompt),
            Message(role="user", content=user_content),
        ]

        # 2. Generate test cases via LLM
        response = await self._llm.generate_chat(
            messages=messages,
            temperature=self._temperature,
            max_tokens=self._max_tokens,
        )
        raw_output = response.content.strip()

        # 3. Parse JSON
        cases = self._parse_cases(raw_output, module)
        if not cases:
            logger.warning("write_test_cases_empty", module=module, raw=raw_output[:200])
            return "LLM 未能生成有效的测试用例，请检查需求文档内容后重试。"

        # 4. Write Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_module = re.sub(r'[\\/:*?"<>|]', "_", module)
        filename = f"{safe_module}_{timestamp}.xlsx"
        file_path = out_dir / filename

        self._write_excel(cases, file_path, module)

        logger.info(
            "test_cases_generated",
            module=module,
            count=len(cases),
            path=str(file_path),
        )
        return (
            f"已生成测试用例 Excel 文件：\n"
            f"路径：{file_path.resolve()}\n"
            f"模块：{module}\n"
            f"用例数量：{len(cases)} 条\n"
            f"（覆盖正向 {sum(1 for c in cases if c.get('type') in ('正向', '功能'))} 条，"
            f"反向/边界/异常 {sum(1 for c in cases if c.get('type') not in ('正向', '功能'))} 条）"
        )

    # ── Parsing ──────────────────────────────────────────────────────────────

    def _parse_cases(self, raw: str, module: str) -> list[dict]:
        """Extract JSON array from LLM output; return normalised case dicts."""
        # Strip markdown fences
        text = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.MULTILINE)
        text = re.sub(r"\s*```$", "", text, flags=re.MULTILINE)
        text = text.strip()

        # Try direct parse
        try:
            data = json.loads(text)
            if isinstance(data, list):
                return [self._normalise(c, module) for c in data if isinstance(c, dict)]
        except json.JSONDecodeError:
            pass

        # Try to find the first JSON array in the text
        match = re.search(r"\[[\s\S]*\]", text)
        if match:
            try:
                data = json.loads(match.group())
                if isinstance(data, list):
                    return [self._normalise(c, module) for c in data if isinstance(c, dict)]
            except json.JSONDecodeError:
                pass

        return []

    @staticmethod
    def _normalise(raw: dict, default_module: str) -> dict:
        """Fill missing fields with safe defaults."""
        return {
            "title": str(raw.get("title", "未命名用例")).strip(),
            "module": str(raw.get("module", default_module)).strip(),
            "precondition": str(raw.get("precondition", "无")).strip(),
            "steps": str(raw.get("steps", "")).strip(),
            "expected": str(raw.get("expected", "")).strip(),
            "priority": str(raw.get("priority", "P1")).strip(),
            "type": str(raw.get("type", "正向")).strip(),
            "notes": str(raw.get("notes", "")).strip(),
        }

    # ── Excel writing ─────────────────────────────────────────────────────────

    def _write_excel(
        self, cases: list[dict], path: Path, module: str
    ) -> None:
        """Write test cases to an xlsx file with header formatting."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError(
                "openpyxl is required. Install with: pip install openpyxl"
            ) from e

        wb = Workbook()
        ws = wb.active
        ws.title = module[:31]  # Excel sheet name limit

        # Header style
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        wrap = Alignment(wrap_text=True, vertical="top")

        # Write headers from config-driven column schema
        for col_idx, (header, width) in enumerate(self._columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        # Write data rows
        prefix = re.sub(r'[^A-Za-z\u4e00-\u9fff]', '', module)[:4].upper() or "TC"
        # Find priority column index (dynamic \u2014 driven by config)
        priority_col = next(
            (i + 1 for i, (name, _) in enumerate(self._columns) if name == "\u4f18\u5148\u7ea7"),
            None,
        )
        for row_idx, case in enumerate(cases, start=2):
            case_id = f"{prefix}-{str(row_idx - 1).zfill(3)}"
            values = [
                case_id,
                case["module"],
                case["title"],
                case["precondition"],
                case["steps"],
                case["expected"],
                case["priority"],
                case["type"],
                case["notes"],
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = wrap

            # Colour-code priority
            if priority_col is not None:
                priority_cell = ws.cell(row=row_idx, column=priority_col)
                priority = case["priority"]
                if priority == "P0":
                    priority_cell.fill = PatternFill("solid", fgColor="FF6B6B")
                    priority_cell.font = Font(bold=True, color="FFFFFF")
                elif priority == "P1":
                    priority_cell.fill = PatternFill("solid", fgColor="FFD93D")
                elif priority == "P2":
                    priority_cell.fill = PatternFill("solid", fgColor="6BCB77")

        wb.save(path)
