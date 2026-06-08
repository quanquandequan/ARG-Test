"""Excel exporter for TestDesignArtifact."""

from __future__ import annotations

import re
from pathlib import Path

from src.application.exporters.common import (
    AUTOMATION_COLUMNS,
    DEFAULT_COLUMNS,
    case_to_dict,
    normalise_generation_mode,
)
from src.domain.artifacts.test_design_artifact import TestDesignArtifact


class ExcelExporter:
    """Export test design artifacts to human-reviewable Excel workbooks."""

    def export(self, artifact: TestDesignArtifact, path: Path) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as e:
            raise RuntimeError(
                "openpyxl is required. Install with: pip install openpyxl"
            ) from e

        wb = Workbook()
        ws = wb.active
        ws.title = artifact.module[:31]

        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        wrap = Alignment(wrap_text=True, vertical="top")
        mode = normalise_generation_mode(artifact.generation_mode)
        columns = AUTOMATION_COLUMNS if mode == "automation" else DEFAULT_COLUMNS

        for col_idx, (header, width) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        prefix = re.sub(r"[^A-Za-z\u4e00-\u9fff]", "", artifact.module)[:4].upper()
        prefix = prefix or "TC"
        priority_col = next(
            (i + 1 for i, (name, _) in enumerate(columns) if name == "优先级"),
            None,
        )

        for row_idx, case in enumerate(artifact.test_cases, start=2):
            row_case = case_to_dict(case)
            case_id = f"{prefix}-{str(row_idx - 1).zfill(3)}"
            values = (
                _automation_row(case_id, row_case)
                if mode == "automation"
                else _manual_row(case_id, row_case)
            )
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = wrap

            if priority_col is not None:
                priority_cell = ws.cell(row=row_idx, column=priority_col)
                priority = row_case["priority"]
                if priority == "P0":
                    priority_cell.fill = PatternFill("solid", fgColor="FF6B6B")
                    priority_cell.font = Font(bold=True, color="FFFFFF")
                elif priority == "P1":
                    priority_cell.fill = PatternFill("solid", fgColor="FFD93D")
                elif priority == "P2":
                    priority_cell.fill = PatternFill("solid", fgColor="6BCB77")

        wb.save(path)


def _manual_row(case_id: str, row_case: dict) -> list[str]:
    return [
        case_id,
        row_case["module"],
        row_case["title"],
        row_case["precondition"],
        row_case["steps"],
        row_case["expected"],
        row_case["priority"],
        row_case["type"],
        row_case["notes"],
    ]


def _automation_row(case_id: str, row_case: dict) -> list[str]:
    return [
        case_id,
        row_case["module"],
        row_case["title"],
        row_case["precondition"],
        row_case.get("data_setup", ""),
        row_case.get("business_name", ""),
        row_case.get("ui_display_name", ""),
        row_case.get("page_route", ""),
        row_case.get("expected_visibility", ""),
        row_case.get("locator_chain", ""),
        row_case.get("anchor_text", ""),
        row_case.get("search_strategy", ""),
        row_case.get("forbidden_locators", ""),
        row_case["steps"],
        row_case["expected"],
        row_case["priority"],
        row_case["type"],
        row_case.get("selectors", ""),
        row_case.get("automation_steps", ""),
        row_case.get("assertions", ""),
        row_case["notes"],
    ]
