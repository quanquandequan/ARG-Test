"""Excel（.xlsx）reader：将结构化单元格数据抽取为可读文本。

每一行会转换为一行文本。第一行视为表头，并作为上下文前缀拼到每个单元格前
（例如 "Status: Open | Priority: High"）。
"""

import uuid
from pathlib import Path

from src.core.exceptions import IngestionError
from src.ingestion.readers.base import BaseReader, Document


class XlsxReader(BaseReader):
    def read(self, path: Path) -> Document:
        if not path.exists():
            raise IngestionError(f"File not found: {path}")

        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise IngestionError(
                "openpyxl is required for .xlsx files. Install with: pip install openpyxl"
            ) from e

        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            raise IngestionError(f"Failed to open Excel file: {path} — {e}") from e

        sheet_names = wb.sheetnames
        all_parts: list[str] = []

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # 第一行作为表头
            headers = [str(h) if h is not None else "" for h in rows[0]]
            data_rows = rows[1:]

            sheet_lines: list[str] = [f"## Sheet: {sheet_name}", ""]

            if not data_rows:
                all_parts.append(f"## Sheet: {sheet_name}\n(empty)")
                continue

            for row_idx, row in enumerate(data_rows, start=2):
                parts: list[str] = []
                for col_idx, value in enumerate(row):
                    header = headers[col_idx] if col_idx < len(headers) else f"Col{col_idx}"
                    cell_text = str(value).strip() if value is not None else ""
                    if cell_text:
                        parts.append(f"{header}: {cell_text}")

                if parts:
                    sheet_lines.append(f"[Row {row_idx}] {' | '.join(parts)}")

            all_parts.append("\n".join(sheet_lines))

        wb.close()

        content = "\n\n".join(all_parts)
        metadata: dict = {
            "sheets": sheet_names,
            "format": "xlsx",
        }

        return Document(
            id=str(uuid.uuid4()),
            source_path=str(path.resolve()),
            content=content,
            metadata=metadata,
        )

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xlsm"]
