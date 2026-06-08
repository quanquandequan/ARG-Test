"""WriteTestCasesTool 测试：需求到 Excel 测试用例生成器。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest

from src.agent.tools.write_test_cases import WriteTestCasesTool
from src.application.artifact_repository import LocalArtifactRepository
from src.application.requirement_services import TestCaseGenerationService
from src.application.workflows.test_case_generation_workflow import (
    TestCaseGenerationWorkflow,
    default_test_case_nodes,
)
from src.llm.types import ChatResponse
from tests.fakes import FakeLLM

# ── Fixture helpers ───────────────────────────────────────────────────────────

_SAMPLE_CASES = [
    {
        "title": "正常登录成功",
        "module": "登录",
        "precondition": "用户已注册",
        "steps": "1. 打开 App\n2. 输入正确账号密码\n3. 点击登录",
        "expected": "登录成功，跳转首页",
        "priority": "P0",
        "type": "正向",
        "notes": "",
    },
    {
        "title": "密码错误登录失败",
        "module": "登录",
        "precondition": "用户已注册",
        "steps": "1. 输入正确账号\n2. 输入错误密码\n3. 点击登录",
        "expected": "提示'密码错误'，停留在登录页",
        "priority": "P0",
        "type": "反向",
        "notes": "",
    },
]

_REQUIREMENT = "用户可以使用账号和密码登录应用。连续失败 5 次需要锁定账号。"

_VALID_IR_JSON = json.dumps({
    "module": "登录",
    "summary": "账号密码登录",
    "actors": [],
    "features": [
        {
            "id": "F001",
            "name": "账号密码登录",
            "description": "用户输入账号密码完成登录",
            "priority": "P0",
            "acceptance_criteria": ["登录成功后进入首页"],
            "test_hints": [],
            "dependencies": [],
        }
    ],
    "business_rules": [],
    "state_machines": [],
    "data_entities": [],
    "out_of_scope": [],
})


def _make_llm_with_cases(cases: list[dict] | None = None) -> FakeLLM:
    payload = json.dumps(cases or _SAMPLE_CASES, ensure_ascii=False)
    return FakeLLM(
        responses=[
            ChatResponse(content=_VALID_IR_JSON, model="fake"),
            ChatResponse(content=payload, model="fake"),
        ]
    )


class _FakeRetrievalEngine:
    async def search(self, **kwargs):
        return []


def _make_service(llm: FakeLLM, output_base: Path) -> TestCaseGenerationService:
    workflow = TestCaseGenerationWorkflow(
        loader=None,
        cleaner=None,
        retrieval_engine=_FakeRetrievalEngine(),
        artifacts=LocalArtifactRepository(base_dir=str(output_base)),
        nodes=default_test_case_nodes(llm),
        default_output_dir=str(output_base),
    )
    return TestCaseGenerationService(workflow=workflow)


def _make_tool_with_cases(
    tmp_path: Path,
    cases: list[dict] | None = None,
) -> tuple[WriteTestCasesTool, FakeLLM]:
    llm = _make_llm_with_cases(cases)
    service = _make_service(llm, tmp_path)
    return WriteTestCasesTool(service=service, output_dir=str(tmp_path)), llm


@pytest.fixture
def tmp_output(tmp_path) -> Path:
    return tmp_path / "test_cases"


@pytest.fixture
def tool(tmp_output) -> WriteTestCasesTool:
    tool, _ = _make_tool_with_cases(tmp_output)
    return tool


# ── Schema / metadata ─────────────────────────────────────────────────────────

def test_tool_name(tool):
    assert tool.name == "write_test_cases"


def test_tool_schema_has_required_requirement(tool):
    schema = tool.to_tool_schema()
    assert "requirement" in schema["parameters"]["required"]
    assert "kb_samples" in schema["parameters"]["properties"]
    assert "module" in schema["parameters"]["properties"]
    assert "output_dir" in schema["parameters"]["properties"]


# ── Happy path ────────────────────────────────────────────────────────────────

async def test_execute_creates_excel_file(tool, tmp_output):
    result = await tool.execute(requirement=_REQUIREMENT, module="登录")
    assert "已生成" in result
    assert ".xlsx" in result
    xlsx_files = list(tmp_output.glob("*.xlsx"))
    assert len(xlsx_files) == 1


async def test_excel_file_has_correct_row_count(tool, tmp_output):
    await tool.execute(requirement=_REQUIREMENT, module="登录")
    from openpyxl import load_workbook
    wb = load_workbook(list(tmp_output.glob("*.xlsx"))[0])
    ws = wb.active
    # 第 1 行为表头，其余行为用例
    data_rows = ws.max_row - 1
    assert data_rows == len(_SAMPLE_CASES)


async def test_excel_file_has_headers(tool, tmp_output):
    await tool.execute(requirement=_REQUIREMENT, module="登录")
    from openpyxl import load_workbook
    wb = load_workbook(list(tmp_output.glob("*.xlsx"))[0])
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "用例标题" in headers
    assert "预期结果" in headers
    assert "优先级" in headers


async def test_execute_automation_mode_has_automation_columns(tmp_path):
    cases = [
        {
            "title": "启动后进入登录页",
            "module": "登录",
            "precondition": "设备已连接，App 已安装",
            "steps": "1. 启动 App\n2. 获取当前屏幕",
            "expected": "页面展示登录按钮",
            "priority": "P0",
            "type": "正向",
            "data_setup": "测试账号已准备",
            "business_name": "登录按钮组件",
            "ui_display_name": "登录",
            "page_route": ["启动 App", "进入登录页"],
            "locator_chain": [{"type": "text", "value": "登录"}],
            "anchor_text": "登录",
            "search_strategy": {"scroll_direction": "none", "max_swipes": 0},
            "expected_visibility": "visible",
            "forbidden_locators": ["登录按钮组件"],
            "selectors": [{"name": "登录按钮", "target_type": "text", "target": "登录"}],
            "automation_steps": [
                {"tool": "device_tool", "action": "connect"},
                {"tool": "device_tool", "action": "launch_app"},
                {"tool": "screen_tool", "action": "get_current_screen"},
            ],
            "assertions": [
                {"tool": "assertion_tool", "action": "assert_text", "text": "登录"}
            ],
            "notes": "",
        }
    ]
    tool, llm = _make_tool_with_cases(tmp_path, cases)

    result = await tool.execute(
        requirement=_REQUIREMENT,
        module="登录",
        generation_mode="automation",
    )

    assert "生成模式：automation" in result
    xlsx_file = next(tmp_path.glob("*automation*.xlsx"))
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_file)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "自动化步骤" in headers
    assert "自动化断言" in headers
    assert "选择器" in headers
    assert "界面展示名" in headers
    assert "禁用定位词" in headers

    json_file = next(tmp_path.glob("*automation*.json"))
    payload = json.loads(json_file.read_text(encoding="utf-8"))
    assert payload["generation_mode"] == "automation"
    assert payload["case_count"] == 1
    assert payload["cases"][0]["ui_display_name"] == "登录"
    assert payload["cases"][0]["page_route"] == ["启动 App", "进入登录页"]
    assert payload["cases"][0]["forbidden_locators"] == ["登录按钮组件"]

    user_msg = next(m for m in llm.last_messages if m.role == "user")
    assert "生成模式：automation" in user_msg.content
    assert "必须区分需求/业务名称和真实 UI 文案" in user_msg.content
    assert "禁止把不可见的需求名" in user_msg.content


async def test_automation_json_keeps_locator_strategy_structured(tmp_path):
    cases = [
        {
            "title": "动画推荐页展示追番表模块",
            "module": "追番表Card",
            "precondition": "当前自然周至少存在1条追番数据，且今天存在更新数据",
            "steps": "1. 启动 App\n2. 点击动画\n3. 点击推荐",
            "expected": "页面展示每日更新模块",
            "priority": "P0",
            "type": "正向",
            "data_setup": "准备本周追番数据和今日更新数据",
            "business_name": "追番表Card",
            "ui_display_name": "每日更新",
            "page_route": ["底部Tab: 动画", "顶部Tab: 推荐"],
            "locator_chain": [
                {"type": "text", "value": "每日更新"},
                {"type": "text", "value": "今"},
                {"type": "anchor", "value": "日漫新作", "relation": "above"},
            ],
            "anchor_text": "日漫新作",
            "search_strategy": {
                "scroll_direction": "down",
                "max_swipes": 8,
                "stop_when_found": ["每日更新", "日漫新作"],
            },
            "expected_visibility": "展示态",
            "forbidden_locators": ["追番表", "追番表Card"],
            "selectors": [{"type": "text", "value": "每日更新"}],
            "automation_steps": [{"tool": "screen_tool", "action": "get_current_screen"}],
            "assertions": [{"tool": "assertion_tool", "action": "assert_text", "text": "每日更新"}],
            "notes": "",
        }
    ]
    tool, _ = _make_tool_with_cases(tmp_path, cases)

    result = await tool.execute(
        requirement="追番表Card 位于日漫新作上方，标题为每日更新。",
        module="追番表Card",
        generation_mode="automation",
    )

    assert "自动化 JSON" in result
    json_file = next(tmp_path.glob("*automation*.json"))
    payload = json.loads(json_file.read_text(encoding="utf-8"))
    case = payload["cases"][0]
    assert case["business_name"] == "追番表Card"
    assert case["ui_display_name"] == "每日更新"
    assert case["expected_visibility"] == "visible"
    assert case["locator_chain"][2] == {
        "type": "anchor",
        "value": "日漫新作",
        "relation": "above",
    }
    assert "追番表Card" in case["forbidden_locators"]


async def test_execute_returns_case_count_in_summary(tool):
    result = await tool.execute(requirement=_REQUIREMENT, module="登录")
    assert f"{len(_SAMPLE_CASES)} 条" in result


async def test_execute_uses_module_in_filename(tool, tmp_output):
    await tool.execute(requirement=_REQUIREMENT, module="登录")
    xlsx_files = list(tmp_output.glob("*.xlsx"))
    assert any("登录" in f.name for f in xlsx_files)


async def test_execute_custom_output_dir(tmp_path):
    custom_dir = tmp_path / "custom_output"
    tool, _ = _make_tool_with_cases(tmp_path / "default")
    result = await tool.execute(
        requirement=_REQUIREMENT,
        output_dir=str(custom_dir),
    )
    assert ".xlsx" in result
    assert list(custom_dir.glob("*.xlsx")), "Excel should be in custom output dir"


async def test_execute_passes_kb_samples_to_llm(tmp_path):
    tool, llm = _make_tool_with_cases(tmp_path)
    kb_sample = "样本用例：正向登录测试"
    await tool.execute(requirement=_REQUIREMENT, kb_samples=kb_sample)
    user_msg = next(m for m in llm.last_messages if m.role == "user")
    assert kb_sample in user_msg.content


async def test_prompt_declares_kb_samples_as_format_only(tmp_path):
    tool, llm = _make_tool_with_cases(tmp_path)
    await tool.execute(
        requirement=_REQUIREMENT,
        kb_samples="历史样本：未登录点击按钮弹出登录弹层",
    )
    sys_msg = next(m for m in llm.last_messages if m.role == "system")
    user_msg = next(m for m in llm.last_messages if m.role == "user")

    assert "需求文档内容】是唯一的业务事实来源" in sys_msg.content
    assert "知识库样本只用于参考 Excel 字段" in sys_msg.content
    assert "样本不是本次需求的事实来源" in user_msg.content


# ── JSON parsing robustness ───────────────────────────────────────────────────

async def test_execute_strips_markdown_fences(tmp_path):
    cases_json = json.dumps(_SAMPLE_CASES, ensure_ascii=False)
    llm = FakeLLM(
        responses=[
            ChatResponse(content=_VALID_IR_JSON, model="fake"),
            ChatResponse(content=f"```json\n{cases_json}\n```", model="fake"),
        ]
    )
    service = _make_service(llm, tmp_path)
    tool = WriteTestCasesTool(service=service, output_dir=str(tmp_path))
    result = await tool.execute(requirement=_REQUIREMENT)
    assert ".xlsx" in result


async def test_execute_handles_json_embedded_in_text(tmp_path):
    cases_json = json.dumps([_SAMPLE_CASES[0]], ensure_ascii=False)
    llm = FakeLLM(
        responses=[
            ChatResponse(content=_VALID_IR_JSON, model="fake"),
            ChatResponse(
                content=f"以下是生成的用例：\n{cases_json}\n希望对你有帮助。",
                model="fake",
            ),
        ]
    )
    service = _make_service(llm, tmp_path)
    tool = WriteTestCasesTool(service=service, output_dir=str(tmp_path))
    result = await tool.execute(requirement=_REQUIREMENT)
    assert ".xlsx" in result


async def test_execute_normalises_missing_fields(tmp_path):
    minimal_cases = [{"title": "最小用例"}]
    llm = FakeLLM(
        responses=[
            ChatResponse(content=_VALID_IR_JSON, model="fake"),
            ChatResponse(content=json.dumps(minimal_cases, ensure_ascii=False), model="fake"),
        ]
    )
    service = _make_service(llm, tmp_path)
    tool = WriteTestCasesTool(service=service, output_dir=str(tmp_path))
    result = await tool.execute(requirement=_REQUIREMENT, module="测试")
    assert ".xlsx" in result


# ── Edge / error cases ────────────────────────────────────────────────────────

async def test_execute_empty_requirement_returns_error(tool):
    result = await tool.execute(requirement="")
    assert "错误" in result


async def test_execute_whitespace_requirement_returns_error(tool):
    result = await tool.execute(requirement="   ")
    assert "错误" in result


async def test_execute_invalid_json_returns_error(tmp_path):
    llm = FakeLLM(
        responses=[
            ChatResponse(content=_VALID_IR_JSON, model="fake"),
            ChatResponse(content="这不是有效的 JSON，也不包含数组。", model="fake"),
        ]
    )
    service = _make_service(llm, tmp_path)
    tool = WriteTestCasesTool(service=service, output_dir=str(tmp_path))
    result = await tool.execute(requirement=_REQUIREMENT)
    assert "LLM 未能生成" in result or "错误" in result


# ── tool_factory integration ──────────────────────────────────────────────────

def test_factory_registers_write_test_cases(
    fake_embedder, fake_vectordb, fake_reranker
):
    from src.agent.tool_factory import build_agent_tools
    from src.retriever.dense_retriever import DenseRetriever
    from src.retriever.retrieval_engine import RetrievalEngine

    dense = DenseRetriever(fake_embedder, fake_vectordb)
    engine = RetrievalEngine(dense_retriever=dense, reranker=fake_reranker)
    service = _make_service(FakeLLM(), Path("/tmp"))

    tools = build_agent_tools(
        engine,
        ["write_test_cases"],
        llm=FakeLLM(),
        test_case_generation_service=service,
    )
    assert len(tools) == 1
    assert tools[0].name == "write_test_cases"


def test_factory_skips_write_test_cases_without_service(
    fake_embedder, fake_vectordb, fake_reranker
):
    from src.agent.tool_factory import build_agent_tools
    from src.retriever.dense_retriever import DenseRetriever
    from src.retriever.retrieval_engine import RetrievalEngine

    dense = DenseRetriever(fake_embedder, fake_vectordb)
    engine = RetrievalEngine(dense_retriever=dense, reranker=fake_reranker)

    tools = build_agent_tools(engine, ["write_test_cases"], llm=FakeLLM())
    assert tools == []


# ── Agent integration ─────────────────────────────────────────────────────────

async def test_agent_orchestrates_kb_search_then_write(
    fake_embedder, fake_vectordb, fake_reranker, tmp_path
):
    """Agent 先调用 knowledge_search，再调用 write_test_cases，返回文件路径。"""
    from src.agent.react_loop import ReActAgent
    from src.agent.tools.search_kb import KnowledgeBaseTool
    from src.llm.types import ToolCall
    from src.retriever.dense_retriever import DenseRetriever
    from src.retriever.retrieval_engine import RetrievalEngine

    dense = DenseRetriever(fake_embedder, fake_vectordb)
    engine = RetrievalEngine(dense_retriever=dense, reranker=fake_reranker)
    kb_tool = KnowledgeBaseTool(engine)

    cases_json = json.dumps(_SAMPLE_CASES, ensure_ascii=False)
    service = _make_service(
        FakeLLM(responses=[_VALID_IR_JSON, cases_json]),
        tmp_path,
    )
    write_tool = WriteTestCasesTool(
        service=service,
        output_dir=str(tmp_path),
    )

    expected_path = str(tmp_path / "登录_20260101.xlsx")
    agent_responses = [
        # 第 1 步：调用 knowledge_search
        ChatResponse(
            content="",
            model="fake",
            stop_reason="tool_use",
            usage={},
            tool_calls=[ToolCall(
                id="tc1", name="knowledge_search",
                arguments={"query": "测试用例 格式"},
            )],
        ),
        # 第 2 步：调用 write_test_cases
        ChatResponse(
            content="",
            model="fake",
            stop_reason="tool_use",
            usage={},
            tool_calls=[ToolCall(
                id="tc2", name="write_test_cases",
                arguments={"requirement": _REQUIREMENT, "module": "登录"},
            )],
        ),
        # 第 3 步：最终答案
        ChatResponse(
            content=f"已生成测试用例文件：\n路径：{expected_path}\n用例数量：2 条",
            model="fake",
            stop_reason="end_turn",
            usage={},
        ),
    ]
    agent_llm = FakeLLM(responses=agent_responses)
    agent = ReActAgent(
        llm=agent_llm,
        tools=[kb_tool, write_tool],
        system_prompt="你是测试助手。",
    )

    result = await agent.run("请为登录功能生成测试用例")
    assert result.iterations == 3
    assert "路径" in result.answer or "已生成" in result.answer
