"""Exporter tests for TestDesignArtifact."""

import json

from openpyxl import load_workbook

from src.application.exporters import ExcelExporter, JsonExporter, MarkdownExporter
from src.domain.artifacts.test_design_artifact import TestDesignArtifact
from src.domain.requirement import Feature, RequirementIR
from src.domain.requirements import GeneratedTestCase
from src.domain.test_design import TestPoint, TestScenario


def _artifact(mode: str = "automation") -> TestDesignArtifact:
    return TestDesignArtifact(
        module="追番表Card",
        generation_mode=mode,
        requirement_ir=RequirementIR(
            module="追番表Card",
            summary="追番表模块",
            features=[
                Feature(id="F001", name="每日更新", description="展示追番更新")
            ],
        ),
        test_points=[TestPoint(id="TP001", title="每日更新展示")],
        scenarios=[TestScenario(id="SC001", title="每日更新展示", point_id="TP001")],
        test_cases=[
            GeneratedTestCase(
                title="动画推荐页展示追番表模块",
                module="追番表Card",
                precondition="当前自然周存在追番数据",
                steps="1. 打开动画推荐页",
                expected="展示每日更新",
                priority="P0",
                case_type="正向",
                data_setup="准备本周追番数据",
                business_name="追番表Card",
                ui_display_name="每日更新",
                page_route=json.dumps(["底部Tab: 动画", "顶部Tab: 推荐"], ensure_ascii=False),
                locator_chain=json.dumps(
                    [{"type": "text", "value": "每日更新"}],
                    ensure_ascii=False,
                ),
                expected_visibility="展示态",
                forbidden_locators=json.dumps(["追番表Card"], ensure_ascii=False),
            )
        ],
    )


def test_excel_exporter_exports_artifact_cases(tmp_path):
    path = tmp_path / "cases.xlsx"
    ExcelExporter().export(_artifact(), path)

    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    assert "界面展示名" in headers
    assert ws.cell(2, 3).value == "动画推荐页展示追番表模块"


def test_json_exporter_preserves_automation_fields(tmp_path):
    path = tmp_path / "cases.json"
    JsonExporter().export(_artifact(), path)

    payload = json.loads(path.read_text(encoding="utf-8"))
    case = payload["cases"][0]
    assert payload["scenario_count"] == 1
    assert case["ui_display_name"] == "每日更新"
    assert case["expected_visibility"] == "visible"
    assert case["forbidden_locators"] == ["追番表Card"]


def test_markdown_exporter_renders_artifact_summary(tmp_path):
    path = tmp_path / "cases.md"
    MarkdownExporter().export(_artifact("manual"), path)

    content = path.read_text(encoding="utf-8")
    assert "测试设计产物：追番表Card" in content
    assert "测试场景：1" in content
